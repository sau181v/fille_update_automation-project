import openpyxl as xl  # xl is alias for openpyxl package
# to add a chart package needed to import

from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']  # accessing the sheet
    cell = sheet['a1']
    for row in range(2, sheet.max_row + 1):  # 1,2,3,4(ignore 1)
        # print(row) # it'll print the values
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        # we can overwrite the new values onto the old ones in the spreadsheet
        # but we will add new column and print the new values there
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        # print(cell.value) # this will print the values...available in the spreadsheet

    # Reference(sheet,
    #           min_row=2,
    #           max_row=sheet.max_row)  # to select a range of values of all rows and column
    # but we only want values of column 4
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    # creating instance of barchart class and store it in object
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


# without the def function for single file.........
'''
# it won't work if there are thousands of code
# because it's relying on a single file 
wb = xl.load_workbook('transactions.xlsx')  # this one
# so for it to work on 1000's of file we have to write a function 
# which will take the input as the diff. file name and do the operation on it.

sheet = wb['Sheet1']  # accessing the sheet
cell = sheet['a1']  # returns and accessin a particular cell
# another approach to access cell
cell = sheet.cell(1, 1)
print(cell.value)  # printing the value of the cell
# # how many rows we have in the sheet
# print(sheet.max_row)
for row in range(2, sheet.max_row + 1):  # 1,2,3,4(ignore 1)
    # print(row) # it'll print the values
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    # we can overwrite the new values onto the old ones in the spreadsheet
    # but we will add new column and print the new values there
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
    # print(cell.value) # this will print the values...available in the spreadsheet

# Reference(sheet,
#           min_row=2,
#           max_row=sheet.max_row)  # to select a range of values of all rows and column
# but we only want values of column 4
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
# creating instance of barchart class and store it in object
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')


wb.save('transactions2.xlsx')  # we are creating new file and saving all the changes with
# the old values in case if some bug/error occurs'''


process_workbook("transactions.xlsx")  # pass the filename here